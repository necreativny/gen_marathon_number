import shutil
import os
from operator import sub, truediv
import smtplib
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase 
from email.mime.multipart import MIMEMultipart 
from email import encoders

from openpyxl import load_workbook
from PIL import ImageFont, ImageDraw, Image
import numpy as np

import config


docx_files = ['Інформація про проведення Забігу 2023.docx', 'Картка учасника Забігу 2023.docx']

wb = load_workbook(filename='membrs.xlsx')
ws = wb.active


server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(config.mail_sender, config.mail_key)

# claer and make new dirs to save results each by distance
for dir in list(config.distance_color) :
	shutil.rmtree(dir, ignore_errors=True)
	os.makedirs(dir)
	for file in docx_files :
		shutil.copy(file, dir)


def draw_text(draw, text, params, font_name, align='tl'):
	font = ImageFont.truetype(font_name, params['font_size'])
	if align == 'tl' :
		point = params['point']
	elif align == 'mid' :
		textbbox = draw.textbbox((0,0), text, font)
		text_size = (
			textbbox[2] - textbbox[0],
			textbbox[3] - textbbox[1]
		)
		text_half_size = tuple(map(truediv,text_size,(2,2)))
		point = tuple(map(sub, params['point'], text_half_size))
	draw.text(point, str(text), font=font, fill=params['color'])

def get_member_data(row):
	res = {}
	for var, col in config.columns.items() :
		res[var] = ws[f'{col}{row}'].value
	return res

def send_mail(receiver, n):
	msg = MIMEMultipart()
	msg['From'] = config.mail_sender
	msg['To'] = receiver
	msg['Subject'] = 'Забіг'
	
	for file in docx_files:
		with open(file, 'rb') as f:
			part = MIMEBase("application", "octet-stream")
			part.set_payload(f.read())
			encoders.encode_base64(part)
			part.add_header('content-disposition', 'attachment', filename = file)
			msg.attach(part)
	with open(n+'.png', 'rb') as f:
		f = MIMEImage(f.read())
		name = n+'.png'
		part.add_header('content-disposition', 'attachment', filename = name)
		msg.attach(f)
	server.sendmail(config.mail_sender, receiver, msg.as_string())

def draw_member_data(memb_data):
	img = Image.open(memb_data['distance'] + '.png')
	draw = ImageDraw.Draw(img)
	for var in list(draw_params) :
		if var == 'n' :
			draw_text(draw, memb_data[var], draw_params[var], config.n_font, align='mid')
		else:
			draw_text(draw, memb_data[var], draw_params[var], config.text_font)
	
	os.chdir(memb_data['distance'])
	img.save(memb_data['n'] + '.png', quality=95)


#if header-- i=1, else i=0
i=0
while True:
	i+=1
	memb_data = get_member_data(i)
	if (memb_data['n']==None) :
		break

	memb_data['n'] = str(memb_data['n'])
	memb_data['distance'] = memb_data['distance'][:-3]#	'2 км' -> '2'
	memb_data['hero_name'] = memb_data['hero_name'].lstrip()+' '
	first_n_end = memb_data['hero_name'].find(' ')
	second_n_end = memb_data['hero_name'][first_n_end+1:].find(' ') + first_n_end+1
	memb_data['hero_name_top'] = memb_data['hero_name'][:first_n_end]#	first name
	memb_data['hero_name_bottom'] = memb_data['hero_name'][first_n_end+1:second_n_end]#	second name
	memb_data['hero_name_bottom'] = memb_data['hero_name_bottom'].upper()
	draw_params = config.draw_params
	draw_params['n']['color'] = config.distance_color[memb_data['distance']]
	
	draw_member_data(memb_data)
	send_mail(memb_data['mail'], memb_data['n'])
	os.chdir(os.path.dirname(os.getcwd()))
